from flask import Flask, send_file
import random
from openpyxl import load_workbook

app = Flask(__name__)

# Load the workbook containing the messages
wb = load_workbook("quotes.xlsx")
ws = wb.active

# Get the total number of rows containing messages
total_messages = ws.max_row

def get_random_message():
    # Select a random row number
    random_row = random.randint(1, total_messages)
    # Get the message from the selected row
    message = ws.cell(row=random_row, column=1).value
    return message

@app.route("/")
def index():
    # Send the HTML file when the root URL is requested
    return send_file("index.html")

@app.route("/get_random_message")
def get_random_message_route():
    # Get a random message
    message = get_random_message()
    alert_message = "Krishna: " + message
    
    return alert_message

if __name__ == "__main__":
    app.run(debug=True)
