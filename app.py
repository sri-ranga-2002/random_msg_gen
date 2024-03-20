import requests
from flask import Flask, send_file
import random
from openpyxl import load_workbook

app = Flask(__name__)

# Load the workbook containing the messages
wb = load_workbook("quotes.xlsx")
ws = wb.active

# Get the total number of rows containing messages
total_messages = ws.max_row

# Account details for SMS sending
api_key = 'NzE2NzcxNGQ0MzY0NmE1MzVhNzQ1MzUzNGE2MjU2Mzc='
api_url = 'https://api.textlocal.in/send/'
recipient_number = '8790110567'  # Replace with the recipient's number
sender_name = 'sri ranga'  # Your sender name

def get_random_message():
    # Select a random row number
    random_row = random.randint(1, total_messages)
    # Get the message from the selected row
    message = ws.cell(row=random_row, column=1).value
    return message

def send_sms(message):
    # Construct the API URL
    url = f"{api_url}?apikey={api_key}&numbers={recipient_number}&message={message}&sender={sender_name}"
    # Send the request
    response = requests.post(url)
    # Print the response (you can handle it as needed)
    print(response.text)

@app.route("/")
def index():
    # Send the HTML file when the root URL is requested
    return send_file("index.html")

@app.route("/get_random_message")
def get_random_message_route():
    # Get a random message
    message = get_random_message()
    alert_message = "Krishna: " + message
    
    # Send SMS
    send_sms(alert_message)
    
    return alert_message

if __name__ == "__main__":
    app.run(debug=True)
