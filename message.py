import openpyxl
import random

def get_random_message(file_path):
    # Load the workbook
    workbook = openpyxl.load_workbook(file_path)
    
    # Select the first worksheet
    sheet = workbook.active
    
    # Get the number of rows in the worksheet
    num_rows = sheet.max_row
    
    # Generate a random row number
    random_row = random.randint(1, num_rows)  # Assuming the messages start from the second row
    
    # Get the message from the random row
    random_message = sheet.cell(row=random_row, column=1).value
    
    return random_message

# Path to the Excel file containing the messages
file_path = "quotes.xlsx"

# Get a random message
random_message = get_random_message(file_path)
print(random_message)
